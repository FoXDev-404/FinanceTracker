from rest_framework import status, generics, viewsets, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.views import TokenObtainPairView
from drf_spectacular.utils import extend_schema, OpenApiParameter
from drf_spectacular.types import OpenApiTypes

from django.conf import settings
from django.db.models import Q, Sum, Count
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import openai
from decimal import Decimal
try:
    import pytesseract
    PYTESSERACT_AVAILABLE = True
except ImportError:
    PYTESSERACT_AVAILABLE = False
from PIL import Image
import json
try:
    import speech_recognition as sr
    SPEECH_RECOGNITION_AVAILABLE = True
except ImportError:
    SPEECH_RECOGNITION_AVAILABLE = False
import base64
import io

# Export libraries
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

from .models import (
    User, Account, Category, Transaction, Budget, Conversation,
    Reminder, Forecast, Anomaly, Tag, RecurringTransaction, SavingsGoal, Notification
)
from .serializers import (
    UserRegistrationSerializer, UserLoginSerializer, UserSerializer,
    AccountSerializer, CategorySerializer, TransactionSerializer, BudgetSerializer,
    TagSerializer, RecurringTransactionSerializer, SavingsGoalSerializer, NotificationSerializer,
    ReminderSerializer, ForecastSerializer, AnomalySerializer
)
from .vision_service import vision_service
from .ai_service import ai_service


class RegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    permission_classes = (AllowAny,)
    serializer_class = UserRegistrationSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            return Response({
                'user': UserSerializer(user).data,
                'message': 'User created successfully'
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class LoginView(TokenObtainPairView):
    permission_classes = (AllowAny,)
    serializer_class = UserLoginSerializer


class UserProfileView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        user = request.user
        serializer = UserSerializer(user)
        return Response(serializer.data)

    def put(self, request):
        user = request.user
        serializer = UserSerializer(user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class LogoutView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        return Response({'message': 'Successfully logged out'})


class AccountViewSet(viewsets.ModelViewSet):
    queryset = Account.objects.all()
    serializer_class = AccountSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return self.queryset.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return self.queryset.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class TagViewSet(viewsets.ModelViewSet):
    queryset = Tag.objects.all()
    serializer_class = TagSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return self.queryset.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class TransactionViewSet(viewsets.ModelViewSet):
    queryset = Transaction.objects.all()
    serializer_class = TransactionSerializer
    permission_classes = [IsAuthenticated]
    ordering = ['-date']
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]

    def get_queryset(self):
        queryset = self.queryset.filter(user=self.request.user)
        params = self.request.query_params
        date_from = params.get('date_from')
        date_to = params.get('date_to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        category = params.get('category')
        if category:
            queryset = queryset.filter(category__category_id=category)
        transaction_type = params.get('transaction_type')
        if transaction_type:
            queryset = queryset.filter(transaction_type=transaction_type)
        min_amount = params.get('min_amount')
        max_amount = params.get('max_amount')
        if min_amount:
            queryset = queryset.filter(amount__gte=min_amount)
        if max_amount:
            queryset = queryset.filter(amount__lte=max_amount)
        tags = params.get('tags')
        if tags:
            tag_ids = [int(t) for t in tags.split(',') if t.isdigit()]
            if tag_ids:
                queryset = queryset.filter(tags__tag_id__in=tag_ids).distinct()
        search = params.get('search')
        if search:
            queryset = queryset.filter(
                Q(note__icontains=search) |
                Q(category__name__icontains=search) |
                Q(account__account_name__icontains=search)
            )
        return queryset.select_related('account', 'category').prefetch_related('tags')

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    def perform_create(self, serializer):
        transaction = serializer.save(user=self.request.user)
        account = transaction.account
        if transaction.transaction_type == 'Income':
            account.balance += transaction.amount
        elif transaction.transaction_type == 'Expense':
            account.balance -= transaction.amount
        account.save()

    def perform_update(self, serializer):
        old_transaction = self.get_object()
        old_account = old_transaction.account
        old_type = old_transaction.transaction_type
        old_amount = old_transaction.amount
        transaction = serializer.save()
        account = transaction.account
        if old_type == 'Income':
            old_account.balance -= old_amount
        elif old_type == 'Expense':
            old_account.balance += old_amount
        if transaction.transaction_type == 'Income':
            account.balance += transaction.amount
        elif transaction.transaction_type == 'Expense':
            account.balance -= transaction.amount
        if old_account != account:
            old_account.save()
        account.save()


class BudgetViewSet(viewsets.ModelViewSet):
    queryset = Budget.objects.all()
    serializer_class = BudgetSerializer
    permission_classes = [IsAuthenticated]
    ordering = ['-month']

    def get_queryset(self):
        return self.queryset.filter(user=self.request.user)

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class RecurringTransactionViewSet(viewsets.ModelViewSet):
    queryset = RecurringTransaction.objects.all()
    serializer_class = RecurringTransactionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return self.queryset.filter(user=self.request.user)

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=False, methods=['post'])
    def generate(self, request):
        user = request.user
        today = timezone.now().date()
        recurring = RecurringTransaction.objects.filter(
            user=user,
            active=True,
            next_due_date__lte=today
        )
        created_count = 0
        for rule in recurring:
            if rule.end_date and rule.end_date < today:
                continue
            Transaction.objects.create(
                user=user,
                account=rule.account,
                category=rule.category,
                amount=rule.amount,
                transaction_type=rule.transaction_type,
                date=today,
                note=f"Recurring: {rule.description or ''}"
            )
            if rule.frequency == 'daily':
                rule.next_due_date += timedelta(days=1)
            elif rule.frequency == 'weekly':
                rule.next_due_date += timedelta(weeks=1)
            elif rule.frequency == 'biweekly':
                rule.next_due_date += timedelta(weeks=2)
            elif rule.frequency == 'monthly':
                rule.next_due_date += relativedelta(months=1)
            elif rule.frequency == 'quarterly':
                rule.next_due_date += relativedelta(months=3)
            elif rule.frequency == 'yearly':
                rule.next_due_date += relativedelta(years=1)
            rule.save()
            created_count += 1
        return Response({"created": created_count})


class SavingsGoalViewSet(viewsets.ModelViewSet):
    queryset = SavingsGoal.objects.all()
    serializer_class = SavingsGoalSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return self.queryset.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=True, methods=['post'])
    def contribute(self, request, pk=None):
        goal = self.get_object()
        amount = Decimal(str(request.data.get('amount', 0)))
        if amount <= 0:
            return Response({"error": "Amount must be greater than zero"}, status=status.HTTP_400_BAD_REQUEST)
        goal.current_amount += amount
        goal.save()
        if goal.current_amount >= goal.target_amount:
            Notification.objects.create(
                user=request.user,
                notification_type='goal_achieved',
                title=f"Goal Achieved: {goal.name}",
                message=f"Congratulations! You've reached your savings goal of ${goal.target_amount} for {goal.name}.",
                data={"goal_id": goal.goal_id, "target": float(goal.target_amount)}
            )
        return Response(SavingsGoalSerializer(goal).data)


class NotificationViewSet(viewsets.ModelViewSet):
    queryset = Notification.objects.all()
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]
    ordering = ['-created_at']

    def get_queryset(self):
        return self.queryset.filter(user=self.request.user)

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        notification = self.get_object()
        notification.read = True
        notification.save()
        return Response(NotificationSerializer(notification).data)

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        count = self.get_queryset().filter(read=False).update(read=True)
        return Response({"marked": count})

    @action(detail=False, methods=['get'])
    def unread_count(self, request):
        count = self.get_queryset().filter(read=False).count()
        return Response({"unread_count": count})


class DashboardStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        now = timezone.now()
        current_month = now.month
        current_year = now.year

        monthly_expenses = []
        for i in range(5, -1, -1):
            month_date = now - relativedelta(months=i)
            month_total = Transaction.objects.filter(
                user=user,
                transaction_type='Expense',
                date__month=month_date.month,
                date__year=month_date.year
            ).aggregate(total=Sum('amount'))['total'] or 0
            monthly_expenses.append({
                'month': month_date.strftime('%b %Y'),
                'amount': float(month_total)
            })

        category_data = Transaction.objects.filter(
            user=user,
            transaction_type='Expense',
            date__month=current_month,
            date__year=current_year
        ).values('category__name', 'category__category_id').annotate(
            total=Sum('amount'),
            count=Count('transaction_id')
        ).order_by('-total')[:8]

        category_pie = [{
            'name': item['category__name'],
            'value': float(item['total']),
            'count': item['count'],
            'id': item['category__category_id']
        } for item in category_data]

        income_vs_expense = []
        for i in range(5, -1, -1):
            month_date = now - relativedelta(months=i)
            income = Transaction.objects.filter(
                user=user,
                transaction_type='Income',
                date__month=month_date.month,
                date__year=month_date.year
            ).aggregate(total=Sum('amount'))['total'] or 0
            expense = Transaction.objects.filter(
                user=user,
                transaction_type='Expense',
                date__month=month_date.month,
                date__year=month_date.year
            ).aggregate(total=Sum('amount'))['total'] or 0
            income_vs_expense.append({
                'month': month_date.strftime('%b %Y'),
                'income': float(income),
                'expense': float(expense),
                'net': float(income - expense)
            })

        current_month_income = Transaction.objects.filter(
            user=user,
            transaction_type='Income',
            date__month=current_month,
            date__year=current_year
        ).aggregate(total=Sum('amount'))['total'] or 0

        current_month_expense = Transaction.objects.filter(
            user=user,
            transaction_type='Expense',
            date__month=current_month,
            date__year=current_year
        ).aggregate(total=Sum('amount'))['total'] or 0

        budgets = Budget.objects.filter(
            user=user,
            month__month=current_month,
            month__year=current_year
        )
        budget_status = []
        for budget in budgets:
            spent = Transaction.objects.filter(
                user=user,
                category=budget.category,
                transaction_type='Expense',
                date__month=current_month,
                date__year=current_year
            ).aggregate(total=Sum('amount'))['total'] or 0
            budget_status.append({
                'category': budget.category.name,
                'budgeted': float(budget.amount),
                'spent': float(spent),
                'remaining': float(budget.amount - spent),
                'percentage': round((float(spent) / float(budget.amount)) * 100, 1) if budget.amount > 0 else 0
            })

        recent_transactions = Transaction.objects.filter(user=user).order_by('-date')[:5]
        recent = TransactionSerializer(recent_transactions, many=True).data

        accounts = Account.objects.filter(user=user)
        account_balances = [{
            'name': acc.account_name,
            'type': acc.account_type,
            'balance': float(acc.balance)
        } for acc in accounts]
        total_balance = sum(acc['balance'] for acc in account_balances)

        goals = SavingsGoal.objects.filter(user=user, active=True)
        savings_goals = [{
            'id': g.goal_id,
            'name': g.name,
            'target': float(g.target_amount),
            'current': float(g.current_amount),
            'progress': g.progress_percentage,
            'icon': g.icon,
            'color': g.color
        } for g in goals]

        return Response({
            'summary': {
                'total_balance': total_balance,
                'monthly_income': float(current_month_income),
                'monthly_expenses': float(current_month_expense),
                'net': float(current_month_income - current_month_expense),
                'savings_rate': round((float(current_month_income - current_month_expense) / float(current_month_income)) * 100, 1) if current_month_income > 0 else 0
            },
            'monthly_expenses_chart': monthly_expenses,
            'category_pie_chart': category_pie,
            'income_vs_expense': income_vs_expense,
            'budget_status': budget_status,
            'recent_transactions': recent,
            'account_balances': account_balances,
            'savings_goals': savings_goals
        })


class BudgetAlertsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        now = timezone.now()
        current_month = now.month
        current_year = now.year

        budgets = Budget.objects.filter(
            user=user,
            month__month=current_month,
            month__year=current_year
        )

        alerts = []
        for budget in budgets:
            spent = Transaction.objects.filter(
                user=user,
                category=budget.category,
                transaction_type='Expense',
                date__month=current_month,
                date__year=current_year
            ).aggregate(total=Sum('amount'))['total'] or 0

            percentage = (float(spent) / float(budget.amount)) * 100 if budget.amount > 0 else 0

            if percentage >= 100:
                alerts.append({
                    'type': 'exceeded',
                    'severity': 'high',
                    'category': budget.category.name,
                    'budgeted': float(budget.amount),
                    'spent': float(spent),
                    'percentage': round(percentage, 1),
                    'message': f'Budget exceeded for {budget.category.name}! Spent ${float(spent):.2f} of ${float(budget.amount):.2f}'
                })
            elif percentage >= 80:
                alerts.append({
                    'type': 'warning',
                    'severity': 'medium',
                    'category': budget.category.name,
                    'budgeted': float(budget.amount),
                    'spent': float(spent),
                    'percentage': round(percentage, 1),
                    'message': f'Budget warning for {budget.category.name}: {round(percentage, 1)}% used'
                })

        return Response({
            'alerts': alerts,
            'has_alerts': len(alerts) > 0
        })


class TransactionExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        export_format = request.query_params.get('format', 'excel').lower()
        queryset = Transaction.objects.filter(user=user)

        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        category = request.query_params.get('category')
        transaction_type = request.query_params.get('transaction_type')

        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        if category:
            queryset = queryset.filter(category__category_id=category)
        if transaction_type:
            queryset = queryset.filter(transaction_type=transaction_type)

        transactions = queryset.select_related('account', 'category').order_by('-date')

        if export_format == 'excel':
            return self._export_excel(transactions)
        elif export_format == 'pdf':
            return self._export_pdf(transactions)
        else:
            return Response({"error": "Invalid format. Use 'excel' or 'pdf'"}, status=status.HTTP_400_BAD_REQUEST)

    def _export_excel(self, transactions):
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_align = Alignment(horizontal="center", vertical="center")

        headers = ["Date", "Account", "Category", "Type", "Amount", "Note"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        for row, t in enumerate(transactions, 2):
            ws.cell(row=row, column=1, value=t.date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=t.account.account_name)
            ws.cell(row=row, column=3, value=t.category.name)
            ws.cell(row=row, column=4, value=t.transaction_type)
            ws.cell(row=row, column=5, value=float(t.amount))
            ws.cell(row=row, column=6, value=t.note or '')

            if t.transaction_type == 'Income':
                ws.cell(row=row, column=4).fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
                ws.cell(row=row, column=4).font = Font(color="FFFFFF")
            else:
                ws.cell(row=row, column=4).fill = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid")
                ws.cell(row=row, column=4).font = Font(color="FFFFFF")

        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 18

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'
        wb.save(response)
        return response

    def _export_pdf(self, transactions):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="transactions.pdf"'

        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=20
        )

        elements.append(Paragraph("Transaction Report", title_style))
        elements.append(Spacer(1, 12))

        total_income = sum(float(t.amount) for t in transactions if t.transaction_type == 'Income')
        total_expense = sum(float(t.amount) for t in transactions if t.transaction_type == 'Expense')

        summary_data = [
            ['Summary', ''],
            ['Total Income', f'${total_income:.2f}'],
            ['Total Expenses', f'${total_expense:.2f}'],
            ['Net', f'${total_income - total_expense:.2f}'],
            ['Total Transactions', str(len(transactions))]
        ]

        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('SPAN', (0, 0), (-1, 0)),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        data = [['Date', 'Account', 'Category', 'Type', 'Amount', 'Note']]
        for t in transactions:
            data.append([
                t.date.strftime('%Y-%m-%d'),
                t.account.account_name,
                t.category.name,
                t.transaction_type,
                f'${float(t.amount):.2f}',
                t.note or ''[:30]
            ])

        table = Table(data, colWidths=[1.2*inch, 1.5*inch, 1.5*inch, 1*inch, 1*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8fafc'), colors.white]),
        ]))
        elements.append(table)

        doc.build(elements)
        return response


class ChatView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user_message = request.data.get('message')
        if not user_message:
            return Response({"error": "Message not provided"}, status=status.HTTP_400_BAD_REQUEST)

        user = request.user
        user_data = self.get_user_financial_data(user)

        conversation_history = Conversation.objects.filter(user=user).order_by('-timestamp')[:10]
        history = []
        for conv in reversed(conversation_history):
            history.append({"role": "user", "content": conv.user_message})
            history.append({"role": "assistant", "content": conv.ai_response})

        system_prompt = f"""You are a helpful AI Finance Assistant for {user.name}. You have access to their financial data and can answer questions about their finances naturally and conversationally.

User's Financial Data Summary:
{user_data}

Guidelines:
- Answer questions based on the provided data.
- Be conversational and helpful.
- If data is missing, suggest adding it.
- For comparisons, calculations, or plans, use the data provided.
- Keep responses concise but informative.
- Provide personalized financial advice based on the user's data.
"""

        messages = [{"role": "system", "content": system_prompt}] + history + [{"role": "user", "content": user_message}]

        structured_response = self.check_structured_queries(user_message, user)
        if structured_response:
            Conversation.objects.create(
                user=user,
                user_message=user_message,
                ai_response=structured_response['text']
            )
            return Response(structured_response)

        try:
            client = openai.OpenAI(api_key=settings.OPENAI_API_KEY)
            response = client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=messages,
                max_tokens=500,
                temperature=0.7
            )
            ai_response = response.choices[0].message.content.strip() if response.choices[0].message.content else "I'm sorry, I couldn't generate a response at the moment."
        except Exception as e:
            ai_response = self.generate_intelligent_fallback_response(user_message, user)

        if not ai_response:
            ai_response = "I'm sorry, I couldn't generate a response at the moment."

        Conversation.objects.create(
            user=user,
            user_message=user_message,
            ai_response=ai_response
        )

        return Response({"response": ai_response}, status=status.HTTP_200_OK)

    def check_structured_queries(self, user_message, user):
        msg_lower = user_message.lower()

        if any(k in msg_lower for k in ['highest spending', 'most spent', 'top expense', 'biggest expense']):
            now = timezone.now()
            current_month = now.month
            current_year = now.year

            top_expense = Transaction.objects.filter(
                user=user,
                transaction_type='Expense',
                date__month=current_month,
                date__year=current_year
            ).order_by('-amount').first()

            top_category = Transaction.objects.filter(
                user=user,
                transaction_type='Expense',
                date__month=current_month,
                date__year=current_year
            ).values('category__name').annotate(total=Sum('amount')).order_by('-total').first()

            text = f"**Highest Spending This Month**\n\n"
            if top_expense:
                text += f"💸 Single highest transaction: **{top_expense.category.name}** - ${float(top_expense.amount):.2f} on {top_expense.date}\n\n"
            if top_category:
                text += f"📊 Top spending category: **{top_category['category__name']}** - ${float(top_category['total']):.2f}\n\n"

            total_spent = Transaction.objects.filter(
                user=user,
                transaction_type='Expense',
                date__month=current_month,
                date__year=current_year
            ).aggregate(total=Sum('amount'))['total'] or 0
            text += f"Total spent this month: **${float(total_spent):.2f}**"

            return {'text': text, 'type': 'highest_spending', 'data': {
                'top_expense': {'category': top_expense.category.name if top_expense else None, 'amount': float(top_expense.amount) if top_expense else 0},
                'top_category': top_category['category__name'] if top_category else None,
                'total_spent': float(total_spent)
            }}

        if any(k in msg_lower for k in ['last 7 days', 'past week', 'weekly report', 'this week']):
            now = timezone.now()
            week_ago = now - timedelta(days=7)

            week_transactions = Transaction.objects.filter(
                user=user,
                date__gte=week_ago.date(),
                date__lte=now.date()
            ).order_by('-date')

            income = week_transactions.filter(transaction_type='Income').aggregate(total=Sum('amount'))['total'] or 0
            expense = week_transactions.filter(transaction_type='Expense').aggregate(total=Sum('amount'))['total'] or 0

            text = f"**Last 7 Days Report**\n\n"
            text += f"📅 Period: {week_ago.strftime('%Y-%m-%d')} to {now.strftime('%Y-%m-%d')}\n\n"
            text += f"💰 Income: **${float(income):.2f}**\n"
            text += f"💸 Expenses: **${float(expense):.2f}**\n"
            text += f"📊 Net: **${float(income - expense):.2f}**\n\n"

            if week_transactions:
                text += f"📝 Transactions ({len(week_transactions)}):\n"
                for t in week_transactions[:10]:
                    emoji = '🟢' if t.transaction_type == 'Income' else '🔴'
                    text += f"{emoji} {t.date}: {t.category.name} - ${float(t.amount):.2f}\n"
            else:
                text += "No transactions in the last 7 days."

            return {'text': text, 'type': 'last_7_days', 'data': {
                'income': float(income),
                'expense': float(expense),
                'net': float(income - expense),
                'transaction_count': len(week_transactions)
            }}

        if any(k in msg_lower for k in ['budget status', 'how are my budgets', 'budget overview']):
            now = timezone.now()
            current_month = now.month
            current_year = now.year

            budgets = Budget.objects.filter(
                user=user,
                month__month=current_month,
                month__year=current_year
            )

            text = f"**Budget Status This Month**\n\n"

            total_budgeted = 0
            total_spent = 0

            for budget in budgets:
                spent = Transaction.objects.filter(
                    user=user,
                    category=budget.category,
                    transaction_type='Expense',
                    date__month=current_month,
                    date__year=current_year
                ).aggregate(total=Sum('amount'))['total'] or 0

                total_budgeted += float(budget.amount)
                total_spent += float(spent)

                pct = (float(spent) / float(budget.amount)) * 100 if budget.amount > 0 else 0
                emoji = '🔴' if pct >= 100 else '🟡' if pct >= 80 else '🟢'
                text += f"{emoji} **{budget.category.name}**: ${float(spent):.2f} / ${float(budget.amount):.2f} ({pct:.1f}%)\n"

            if budgets:
                text += f"\n💰 Total Budgeted: ${total_budgeted:.2f}\n"
                text += f"💸 Total Spent: ${total_spent:.2f}\n"
                text += f"📊 Overall: {((total_spent / total_budgeted) * 100):.1f}% used"
            else:
                text += "No budgets set for this month."

            return {'text': text, 'type': 'budget_status', 'data': {
                'total_budgeted': total_budgeted,
                'total_spent': total_spent,
                'percentage': round((total_spent / total_budgeted) * 100, 1) if total_budgeted > 0 else 0
            }}

        if any(k in msg_lower for k in ['savings progress', 'goals', 'savings goals', 'goal progress']):
            goals = SavingsGoal.objects.filter(user=user, active=True)

            text = f"**Savings Goals Progress**\n\n"

            for goal in goals:
                progress = goal.progress_percentage
                emoji = '🎉' if progress >= 100 else '🔥' if progress >= 75 else '💪' if progress >= 50 else '📈'
                text += f"{emoji} **{goal.name}**: ${float(goal.current_amount):.2f} / ${float(goal.target_amount):.2f} ({progress:.1f}%)\n"
                text += f"   Remaining: ${goal.remaining_amount:.2f}\n\n"

            if not goals:
                text += "No active savings goals. Create one to start tracking your progress!"

            return {'text': text, 'type': 'savings_progress', 'data': [
                {
                    'name': g.name,
                    'target': float(g.target_amount),
                    'current': float(g.current_amount),
                    'progress': g.progress_percentage
                } for g in goals
            ]}

        if any(k in msg_lower for k in ['monthly report', 'this month summary', 'month summary']):
            now = timezone.now()
            current_month = now.month
            current_year = now.year

            income = Transaction.objects.filter(
                user=user,
                transaction_type='Income',
                date__month=current_month,
                date__year=current_year
            ).aggregate(total=Sum('amount'))['total'] or 0

            expense = Transaction.objects.filter(
                user=user,
                transaction_type='Expense',
                date__month=current_month,
                date__year=current_year
            ).aggregate(total=Sum('amount'))['total'] or 0

            top_categories = Transaction.objects.filter(
                user=user,
                transaction_type='Expense',
                date__month=current_month,
                date__year=current_year
            ).values('category__name').annotate(total=Sum('amount')).order_by('-total')[:5]

            text = f"**Monthly Report: {now.strftime('%B %Y')}**\n\n"
            text += f"💰 Total Income: **${float(income):.2f}**\n"
            text += f"💸 Total Expenses: **${float(expense):.2f}**\n"
            text += f"📊 Net: **${float(income - expense):.2f}**\n\n"

            if top_categories:
                text += f"📈 Top Spending Categories:\n"
                for cat in top_categories:
                    text += f"• {cat['category__name']}: ${float(cat['total']):.2f}\n"

            return {'text': text, 'type': 'monthly_report', 'data': {
                'income': float(income),
                'expense': float(expense),
                'net': float(income - expense),
                'top_categories': [
                    {'name': c['category__name'], 'amount': float(c['total'])}
                    for c in top_categories
                ]
            }}

        return None

    def get_user_financial_data(self, user):
        data_parts = []

        accounts = Account.objects.filter(user=user)
        if accounts:
            account_summary = "\n".join([f"- {acc.account_name} ({acc.account_type}): ${acc.balance}" for acc in accounts])
            total_balance = sum(acc.balance for acc in accounts)
            data_parts.append(f"Accounts:\n{account_summary}\nTotal Balance: ${total_balance}")
        else:
            data_parts.append("Accounts: None")

        categories = Category.objects.filter(user=user)
        if categories:
            category_summary = "\n".join([f"- {cat.name} ({cat.type})" for cat in categories])
            data_parts.append(f"Categories:\n{category_summary}")
        else:
            data_parts.append("Categories: None")

        transactions = Transaction.objects.filter(user=user).order_by('-date')[:20]
        if transactions:
            transaction_summary = "\n".join([
                                f"- {t.date}: {t.category.name} - {t.transaction_type} ${t.amount} ({t.note or 'No note'})"
                for t in transactions
            ])
            data_parts.append(f"Recent Transactions:\n{transaction_summary}")
        else:
            data_parts.append("Recent Transactions: None")

        current_month = datetime.now().month
        current_year = datetime.now().year
        budgets = Budget.objects.filter(user=user, month__month=current_month, month__year=current_year)
        if budgets:
            budget_summary = "\n".join([f"- {b.category.name}: ${b.amount}" for b in budgets])
            total_budget = sum(b.amount for b in budgets)
            data_parts.append(f"Current Month Budgets:\n{budget_summary}\nTotal Budget: ${total_budget}")
        else:
            data_parts.append("Current Month Budgets: None")

        spending_by_category = Transaction.objects.filter(
            user=user,
            transaction_type='Expense',
            date__month=current_month,
            date__year=current_year
        ).values('category__name').annotate(total=Sum('amount')).order_by('-total')
        if spending_by_category:
            spending_summary = "\n".join([f"- {item['category__name']}: ${item['total']}" for item in spending_by_category])
            data_parts.append(f"Current Month Spending by Category:\n{spending_summary}")
        else:
            data_parts.append("Current Month Spending by Category: None")

        return "\n\n".join(data_parts)

    def generate_intelligent_fallback_response(self, user_message, user):
        user_message_lower = user_message.lower()
        now = datetime.now()
        current_month = now.month
        current_year = now.year
        accounts = Account.objects.filter(user=user)
        transactions = Transaction.objects.filter(user=user)

        if any(k in user_message_lower for k in ['hello', 'hi', 'hey']):
            return f"Hello {user.name}! I'm your AI Finance Assistant. What would you like to know?"

        if any(k in user_message_lower for k in ['balance', 'how much money', 'total balance']):
            if accounts:
                total_balance = sum(float(acc.balance) for acc in accounts)
                return f"Your total balance is ${total_balance:.2f}."
            return "You don't have any accounts set up yet."

        if any(k in user_message_lower for k in ['spend', 'spent', 'expense', 'how much did I spend']):
            expenses = transactions.filter(transaction_type='Expense', date__month=current_month, date__year=current_year)
            if expenses:
                total_spent = sum(float(exp.amount) for exp in expenses)
                return f"This month, you've spent ${total_spent:.2f}."
            return "You haven't recorded any expenses this month yet."

        if any(k in user_message_lower for k in ['budget', 'budgets']):
            budgets = Budget.objects.filter(user=user, month__month=current_month, month__year=current_year)
            if budgets:
                total_budget = sum(float(b.amount) for b in budgets)
                return f"Your total budget for this month is ${total_budget:.2f}."
            return "You don't have any budgets set for this month."

        if any(k in user_message_lower for k in ['income', 'earn', 'salary']):
            monthly_income = transactions.filter(
                transaction_type='Income',
                date__month=current_month,
                date__year=current_year
            ).aggregate(total=Sum('amount'))['total'] or 0
            if monthly_income:
                return f"This month, you've earned ${float(monthly_income):.2f} in income."
            return "You haven't recorded any income this month yet."

        return "I'm here to help with your finances! Ask me about your balance, spending, budgets, or transactions."


class ReminderViewSet(viewsets.ModelViewSet):
    queryset = Reminder.objects.all()
    serializer_class = ReminderSerializer
    permission_classes = [IsAuthenticated]
    ordering = ['due_date']

    def get_queryset(self):
        return self.queryset.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=False, methods=['get'])
    def upcoming(self, request):
        today = timezone.now().date()
        week_later = today + timedelta(days=7)
        reminders = self.get_queryset().filter(
            active=True,
            due_date__gte=today,
            due_date__lte=week_later
        )
        data = [{
            'reminder_id': r.reminder_id,
            'title': r.title,
            'description': r.description,
            'due_date': r.due_date,
            'amount': float(r.amount) if r.amount else None,
            'reminder_type': r.reminder_type,
            'recurring': r.recurring
        } for r in reminders]
        return Response(data)


class ForecastViewSet(viewsets.ModelViewSet):
    queryset = Forecast.objects.all()
    permission_classes = [IsAuthenticated]
    ordering = ['-period_start']

    def get_queryset(self):
        return self.queryset.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=False, methods=['post'])
    def generate(self, request):
        forecast_type = request.data.get('forecast_type', 'expense')
        periods = int(request.data.get('periods', 3))
        result = ai_service.generate_forecast(request.user, forecast_type, periods)
        return Response(result)


class AnomalyViewSet(viewsets.ModelViewSet):
    queryset = Anomaly.objects.all()
    permission_classes = [IsAuthenticated]
    ordering = ['-created_at']

    def get_queryset(self):
        return self.queryset.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=False, methods=['post'])
    def detect(self, request):
        anomalies = ai_service.detect_anomalies(request.user)
        created = []
        for anomaly_data in anomalies:
            anomaly = Anomaly.objects.create(
                user=request.user,
                anomaly_type=anomaly_data['type'],
                severity=anomaly_data.get('severity', 'medium'),
                description=f"Unusual spending detected: ${anomaly_data['amount']:.2f} on {anomaly_data['date']}",
                confidence_score=0.85
            )
            created.append({
                'anomaly_id': anomaly.anomaly_id,
                'type': anomaly.anomaly_type,
                'severity': anomaly.severity,
                'description': anomaly.description
            })
        return Response({'detected': len(created), 'anomalies': created})

    @action(detail=True, methods=['post'])
    def resolve(self, request, pk=None):
        anomaly = self.get_object()
        anomaly.resolved = True
        anomaly.resolved_at = timezone.now()
        anomaly.save()
        return Response({'status': 'resolved'})


class VoiceInputView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        audio_text = request.data.get('audio_text')
        if not audio_text:
            return Response({"error": "No audio text provided"}, status=status.HTTP_400_BAD_REQUEST)

        result = ai_service.process_voice_input(audio_text, request.user)
        return Response(result)


class ReceiptProcessView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        image = request.FILES.get('image')
        if not image:
            return Response({"error": "No image provided"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Use pytesseract for OCR if available, otherwise fall back to raw bytes
            img = Image.open(image)
            if PYTESSERACT_AVAILABLE:
                text = pytesseract.image_to_string(img)
            else:
                return Response({"error": "OCR service not available on this deployment"}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

            # Use AI to parse receipt
            prompt = f"""Parse this receipt text and extract:
- vendor/store name
- date
- total amount
- items (if visible)
- category

Receipt text:
{text}

Return as JSON with keys: vendor, date (YYYY-MM-DD), amount, items (array), category"""

            client = openai.OpenAI(api_key=settings.OPENAI_API_KEY)
            response = client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=300
            )

            result = json.loads(response.choices[0].message.content.strip())
            result['raw_text'] = text
            return Response(result)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
